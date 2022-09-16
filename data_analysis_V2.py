import numpy as np
from openpyxl import Workbook, load_workbook 
from openpyxl.chart import LineChart, Reference
import cv2
from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
from scipy import signal
from matplotlib.animation import PillowWriter
import time
from matplotlib.animation import FFMpegWriter
plt.rcParams['animation.ffmpeg_path'] = 'C:\\Users\\anisu\\OneDrive\\Desktop\\ffmpeg-2022-07-24-git-39a538f430-full_build\\bin\\ffmpeg.exe'



def filter_signal(fc, fs, filter_order, data):
	w = fc / (fs/2)
	b, a = signal.butter(filter_order, w, 'low', analog=False)
	filtered_signal = signal.filtfilt(b, a, data)
	return filtered_signal

def predict(times, velocity_data, acceleration_data, actuator_threshold):
	dI = velocity_data
	dI2 = acceleration_data
	streaks = {}
	positive_count = 0
	i = 0
	N = len(times)
	while i < N:
		if dI2[i] == min(dI2):
			rain_time = times[i]
			break
		i += 1

	while i < N-1:
		if dI[i] == max(dI[i:]):
			actuator_time = times[i]
			break
		i += 1
		# if dI[i] > 0:
		# 	positive_count += 1
		# else:
		# 	streaks[positive_count] = times[i - positive_count + 1]
		# 	positive_count = 0
		# # if positive_count == actuator_threshold:
		# # 	actuator_time = times[i]
		# # 	break
		# i += 1

	# print(streaks)
	# print(list(sorted(streaks.keys(), reverse=True)))
	# original_keys_list = list(streaks.keys())
	# key_list= []
	# for key in original_keys_list:
	# 	#if streaks[key] < streaks[max(original_keys_list)]:
	# 	key_list.append(key) #
	# actuator_time = streaks[max(key_list)]
	return [rain_time, actuator_time]
#==================================================================================================================================================


path = Path.cwd() / '21A2' /'Output.xlsx'
# cap = cv2.VideoCapture('./21A2.mp4')
# fps = cap.get(cv2.CAP_PROP_FPS)      # OpenCV2 version 2 used "CV_CAP_PROP_FPS"
# frame_count = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
# duration = frame_count/fps

data = pd.read_excel(path)
intensity_data = pd.DataFrame(data, columns = ['Intensity']).to_numpy()
raw_intensities = []
raw_intensities_derivative = [0]
raw_intensities_derivative2 = [0]
max_raw = max(intensity_data)
N = len(intensity_data)
duration = N/17
T = duration / N
fs = 1 / T
times = np.linspace(0, duration, N, endpoint=True)
i = 0
for val in intensity_data:
	raw_intensities.append(float(val / max_raw * 100))
	if i > 0:
		raw_intensities_derivative.append((raw_intensities[i] - raw_intensities[i-1]) / (times[i] - times[i-1]))
		raw_intensities_derivative2.append((raw_intensities_derivative[i] - raw_intensities_derivative[i-1]) / (times[i] - times[i-1]))
	i += 1

fc = 0.9
w = fc / (fs/2)
b, a = signal.butter(1, w, 'low', analog=False)
org_filtered_intensities = signal.filtfilt(b, a, raw_intensities)
max_val = max(org_filtered_intensities)
filtered_intensities = []
for i in range(N):
	filtered_intensities.append(org_filtered_intensities[i] / max_val * 100)

dI = [0]
dI2 = [0]
cleaning_array = []
cleaning_time = []
i = 1
wb = Workbook()
curr_sheet = wb.create_sheet()
curr_sheet.cell(1, 1, 'Time (s)')
curr_sheet.cell(1, 2, 'Intensity (%)')
curr_sheet.cell(1, 3, 'Intensity Improvement Rate (%/s)')
curr_sheet.cell(1, 4, 'Intensity Improvement Acceleration (%/s/s)')
negative_count = 0
positive_count = 0
rain_threshold = 25
actuator_threshold = 20

while i < N:
	val = (filtered_intensities[i] - filtered_intensities[i-1]) / (times[i] - times[i-1])
	dI.append(val)
	val2 = (val - dI[i-1]) / (times[i] - times[i-1])
	dI2.append(val2)
	curr_sheet.cell(i+1, 1, round(times[i], 3))
	curr_sheet.cell(i+1, 2, filtered_intensities[i])
	curr_sheet.cell(i+1, 3, val)
	curr_sheet.cell(i+1, 4, val2)
	i += 1

dI = filter_signal(0.9, fs, 2, dI)
dI2 = filter_signal(0.9, fs, 2, dI2)
ws = curr_sheet

chart = LineChart()
vals = Reference(ws, min_col = 3, min_row = 1, max_col = 3, max_row = i)
x_values = Reference(ws, min_col = 1, min_row = 2, max_col = 1, max_row = i)
chart.set_categories(x_values)
chart.add_data(vals, titles_from_data=True)
chart.x_axis.title = 'Time (s)'
chart.y_axis.title = 'Intensity Improvement Rate (%/s)'
chart.title = 'LiDAR Intensity and Improvement Rate vs Time'
chart.height = 15
chart.width = 30

chart2 = LineChart()
vals2 = Reference(ws, min_col = 2, min_row = 1, max_col = 2, max_row = i)
x_values2 = Reference(ws, min_col = 1, min_row = 2, max_col = 1, max_row = i)
chart2.add_data(vals2, titles_from_data = True)
chart2.y_axis.axId = 200
chart2.set_categories(x_values2)
chart2.y_axis.title = 'LiDAR Intensity (%)'
chart2.y_axis.majorGridlines = None

chart.y_axis.crosses = 'max'
chart += chart2
ws.add_chart(chart, 'H4')

wb.remove(wb['Sheet'])
wb.save('Analysis Results.xlsx')


rain_time, actuator_time = predict(times, dI, dI2, 20)
#actuator_time = streaks[max(list(sorted(streaks.keys(), reverse=True)))]

print('LiDAR rain and actuator prediciton: ' + str([round(rain_time, 2), round(actuator_time, 2)]))



OLA_data = pd.read_excel(Path(r'C:\Users\anisu\OneDrive\Desktop\lidarvideos\21a22_\Output_Edge.xlsx'))
cleanEff = pd.DataFrame(OLA_data, columns = ['Time', 'ObscurationNorm1']).to_numpy()
OLA_times = []
OLA_eff_raw = []

for arr in cleanEff:
	OLA_times.append(arr[0])
	OLA_eff_raw.append(arr[1])

OLA_eff_filt = filter_signal(0.9, 30, 1, OLA_eff_raw)
maxOLA = max(OLA_eff_filt)

for i in range(len(OLA_eff_filt)):
	OLA_eff_filt[i] = OLA_eff_filt[i] / maxOLA * 100

OLA_speed = [0]
OLA_acc = [0]
for i in range(1, len(OLA_eff_filt)):
	val = (OLA_eff_filt[i] - OLA_eff_filt[i-1]) / (OLA_times[i] -  OLA_times[i-1])
	val2 = (val - OLA_speed[i-1]) / (OLA_times[i] -  OLA_times[i-1])
	OLA_speed.append(val)
	OLA_acc.append(val2)

OLA_speed_filtered = filter_signal(0.9, 30, 2, OLA_speed)
OLA_acc_filtered = filter_signal(0.7, 30, 4, OLA_acc)

OLA_rain, OLA_actuator = predict(OLA_times, OLA_speed, OLA_speed, 20)
print('OLA rain and actuator prediciton: ' + str([round(OLA_rain, 2), round(OLA_actuator, 2)]))


lidar_cleaining_times = times - actuator_time
time_list = []
for time in times:
	time_list.append(time)
act_index = time_list.index(actuator_time)

min_val = min(OLA_times, key = lambda x:abs(x-actuator_time))
OLA_cleaning_times = OLA_times - min_val
OLA_index = OLA_times.index(min_val) - 10

show = False
if show:
	fig1 = plt.figure(1)
	plt.plot(OLA_times[:1029], OLA_eff_filt[:1029], label = 'OLA')
	plt.plot(times, filtered_intensities, label = 'LiDAR')
	plt.xlabel('Time (s)')
	plt.ylabel('Normalized Values')
	plt.legend(loc = 'best')
	plt.title('LiDAR and OLA Data')

	fig2 = plt.figure(2)
	plt.plot(OLA_cleaning_times[OLA_index:1029], OLA_eff_filt[OLA_index:1029], label = 'OLA')
	plt.plot(lidar_cleaining_times[act_index:1029], filtered_intensities[act_index:1029], label = 'LiDAR')
	plt.xlabel('Time (s)')
	plt.ylabel('Normalized Data')
	plt.title('Data Values Over Cleaning Interval')
	plt.legend(loc = 'best')
	
	fig3 = plt.figure(3)
	plt.plot(OLA_times[:1029], OLA_speed[:1029], label = 'OLA')
	plt.plot(times, dI, label = 'LiDAR')
	plt.xlabel('Time (s)')
	plt.ylabel('Normalized Speed Improvement (%/s)')
	plt.legend(loc = 'best')
	plt.title('OLA Obscuration and LiDAR Intensity Improvement Rate vs Time')

	fig4 = plt.figure(4)
	plt.plot(times, raw_intensities_derivative, '-r',label='Raw Data', linewidth = 0.5)
	plt.plot(times, dI, '-c', label='Filtered Data', linewidth = 3)
	plt.xlabel('Time (s)')
	plt.ylabel('Intensity Improvement Rate (%/s)')
	plt.title('Filtered and Unfiltered Inensity Improvement Rate')
	plt.legend(loc='best')

	fig5 = plt.figure(5)
	plt.plot(times, raw_intensities, '-r', label='Raw Data', linewidth = 0.5)
	plt.plot(times, org_filtered_intensities, '-c', label = 'Filtered Data', linewidth = 3)
	plt.xlabel('Time (s)')
	plt.ylabel('Intensity Improvement Data (%)')
	plt.title('Filtered and Unfiltered Inensity')
	plt.legend(loc='best')
	
	fig6 = plt.figure(6)
	plt.plot(OLA_times, OLA_speed, '-r', label='Raw Data', linewidth = 0.5)
	plt.plot(OLA_times, OLA_speed_filtered, '-c', label='Filtered Data', linewidth = 3)
	plt.xlabel('Time (s)')
	plt.ylabel(r'Effectiveness Improvement Rate $\frac{\%}{s}$')
	plt.title('Filtered and Unfiltered Inensity Improvement Rate')
	plt.legend(loc='best')
	
	fig7 = plt.figure(7)
	plt.plot(times, dI2, '-r', label='Filtered Data', linewidth = 3)
	plt.xlabel('Time (s)')
	plt.ylabel(r'Intensity Improvement Acceleration $\frac{\%}{s^2}$')
	plt.title('Filtered and Inensity Improvement Acceleration')
	plt.legend(loc='best')

	plt.show()

	# plt.plot(OLA_times, OLA_eff_raw, '-r',linewidth=2)
	# plt.xlabel('Time (s)')
	# plt.ylabel(r'OLA Cleaning Effectiveness (%)')
	# plt.title('OLA Cleaning Effectiveness vs Time')
	# plt.legend(loc='best')
	# plt.show()


# for i in range(1,15):
# 	OLA_acc_test = filter_signal(0.75, 30, i, OLA_acc)
# 	plt.plot(OLA_times, OLA_acc, '-b', linewidth= 0.75, label = 'Raw Data') 
# 	plt.plot(OLA_times, OLA_acc_test, '-k', linewidth=3, label='Filtered Data')
# 	plt.legend(loc='best')
# 	plt.show(block=False)
# 	plt.pause(0.5)
# 	plt.clf()


# fig = plt.figure()
# metadata = dict(title = 'movie', artist='Me')
# writer = FFMpegWriter(fps = 2, metadata=metadata)

# with writer.saving(fig, 'filter_effect.mp4', 200):
# 	for i in range(1, 16):
# 		filtered_data = filter_signal(0.9, fs, i, raw_intensities)
# 		l = plt.plot(times, raw_intensities, '-r', label='Raw Data', linewidth=1)
# 		plt.xlim(0, max(times))
# 		plt.ylim(0, 110)
# 		m = plt.plot(times, filtered_data, '-k', label='Filter Order: ' + str(i), linewidth=2)
# 		plt.legend(loc='best')
# 		writer.grab_frame()
# 		plt.clf()

# w = np.linspace(0.99, 0.1, 15)
# fig = plt.figure()
# with writer.saving(fig, 'frequency_effect.mp4', 200):
# 	for f in w:
# 		filtered_data = filter_signal(f, fs, 4, raw_intensities)
# 		l = plt.plot(times, raw_intensities, '-r', label='Raw Data', linewidth=1)
# 		plt.xlim(0, max(times))
# 		plt.ylim(0, 110)
# 		m = plt.plot(times, filtered_data, '-k', label=r'$\frac{\omega}{\omega_n}$ = ' + str(round(f,2)), linewidth=2)
# 		plt.legend(loc='best')
# 		writer.grab_frame()
# 		plt.clf()



# fig = plt.figure(1)
# noise = np.random.normal(loc=0, scale=1, size=N,)
# plt.plot(noise, '-r')

# fig2 = plt.figure(2)
# plt.plot(times, np.asarray(raw_intensities_derivative2) / max(raw_intensities_derivative2), '-r')
# plt.show()

# print(raw_intensities_derivative2.index(min(raw_intensities_derivative2)))
# print(raw_intensities_derivative2.index(max(raw_intensities_derivative2)))
# print(filtered_intensities[180])
# print(filtered_intensities[200])

def differentiate(*, ydata, xdata, order = 1):
	"""Take in a set of data and numerically differentiate n times"""

	final_lists = []
	final_lists.append(ydata)
	count = 0
	N = order + 1
	while count < N:
		fx = final_lists[count]
		ddx = [0]	
		for i in range(1, len(ydata)):
			ddx.append((fx[i] - fx[i-1]) / (xdata[i] - xdata[i-1]))
		final_lists.append(ddx)
		count += 1
	return final_lists[1:N]


